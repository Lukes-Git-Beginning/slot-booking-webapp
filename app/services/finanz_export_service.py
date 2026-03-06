# -*- coding: utf-8 -*-
"""
Finanzberatung Export Service

Generates Excel and PDF exports for financial advisory sessions.

Excel (openpyxl): 6 sheets
1. Uebersicht — Customer, advisor, date, overall score, category Ampel
2. Vertraege — All contracts with type, company, premium, status
3. Detailfelder — ALL extracted fields with source references
4. Kosten — Aggregated premiums per category + total
5. Optimierung — Scorecard assessments + missing fields + recommendations
6. Rohdaten — All FinanzExtractedData records with technical details

PDF (reportlab): Structured document with cover, Ampel, contracts, actions
"""

import io
import logging
from datetime import datetime

from app.config.finanz_checklist import (
    CONTRACT_TYPES, CHECKLIST_CATEGORIES, PRIORITY_LABELS,
    get_fields_for_type, compute_completeness,
)
from app.models import get_db_session
from app.models.finanzberatung import (
    FinanzSession, FinanzDocument, FinanzExtractedData, FinanzScorecard,
    ScorecardCategory, TrafficLight, DocumentStatus,
)

logger = logging.getLogger(__name__)


# Optional dependency checks
try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import reportlab  # noqa: F401
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class FinanzExportService:
    """Generates Excel and PDF exports for Finanzberatung sessions."""

    def export_excel(self, session_id: int) -> io.BytesIO:
        """
        Generate Excel export with 6 sheets.

        Args:
            session_id: Session ID

        Returns:
            BytesIO buffer containing the Excel file
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl nicht installiert — Excel-Export nicht verfügbar")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        db = get_db_session()
        try:
            session = db.query(FinanzSession).filter(
                FinanzSession.id == session_id
            ).first()
            if session is None:
                raise ValueError(f"Session {session_id} not found")

            docs = db.query(FinanzDocument).filter(
                FinanzDocument.session_id == session_id,
            ).order_by(FinanzDocument.created_at).all()

            scorecards = db.query(FinanzScorecard).filter(
                FinanzScorecard.session_id == session_id
            ).all()

            wb = Workbook()

            # Styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="294c5d", end_color="294c5d", fill_type="solid")
            gold_fill = PatternFill(start_color="d4af6a", end_color="d4af6a", fill_type="solid")
            green_fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
            yellow_fill = PatternFill(start_color="eab308", end_color="eab308", fill_type="solid")
            red_fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )

            def style_header(ws, row=1):
                for cell in ws[row]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

            def rating_fill(rating):
                if rating == TrafficLight.GREEN:
                    return green_fill
                elif rating == TrafficLight.YELLOW:
                    return yellow_fill
                elif rating == TrafficLight.RED:
                    return red_fill
                return None

            # --- Sheet 1: Uebersicht ---
            ws1 = wb.active
            ws1.title = "Uebersicht"
            ws1.append(["Finanzberatung - Uebersicht"])
            ws1.merge_cells("A1:D1")
            ws1['A1'].font = Font(bold=True, size=16, color="294c5d")
            ws1.append([])
            ws1.append(["Kunde", session.customer_name])
            ws1.append(["Opener", session.opener_username])
            ws1.append(["Closer", session.closer_username or "Nicht zugewiesen"])
            ws1.append(["Beratungsart", session.session_type])
            ws1.append(["Termin", session.appointment_date.strftime('%d.%m.%Y') if session.appointment_date else ""])
            ws1.append(["Status", session.status])
            ws1.append(["Exportiert am", datetime.now().strftime('%d.%m.%Y %H:%M')])
            ws1.append([])

            # Scorecard summary
            ws1.append(["Kategorie", "Bewertung", "Zusammenfassung"])
            style_header(ws1, ws1.max_row)
            cat_labels = {
                ScorecardCategory.ALTERSVORSORGE: "Altersvorsorge",
                ScorecardCategory.ABSICHERUNG: "Absicherung",
                ScorecardCategory.VERMOEGEN_KOSTEN: "Vermoegen & Kosten",
                ScorecardCategory.STEUEROPTIMIERUNG: "Steueroptimierung",
            }
            rating_labels = {
                TrafficLight.GREEN: "Gruen",
                TrafficLight.YELLOW: "Gelb",
                TrafficLight.RED: "Rot",
            }
            for sc in scorecards:
                if sc.is_overall:
                    continue
                label = cat_labels.get(sc.category, sc.category)
                r_label = rating_labels.get(sc.rating, sc.rating)
                row_num = ws1.max_row + 1
                ws1.append([label, r_label, sc.assessment or ""])
                fill = rating_fill(sc.rating)
                if fill:
                    ws1.cell(row=row_num, column=2).fill = fill

            # Overall
            overall = next((sc for sc in scorecards if sc.is_overall), None)
            if overall:
                ws1.append([])
                ws1.append(["Gesamtbewertung", rating_labels.get(overall.rating, ""), overall.assessment or ""])

            ws1.column_dimensions['A'].width = 25
            ws1.column_dimensions['B'].width = 20
            ws1.column_dimensions['C'].width = 60

            # --- Sheet 2: Vertraege ---
            ws2 = wb.create_sheet("Vertraege")
            ws2.append(["Vertragsart", "Gesellschaft", "Beitrag", "Status", "Dokument", "Confidence"])
            style_header(ws2)

            for doc in docs:
                if not doc.document_type or doc.status != DocumentStatus.ANALYZED:
                    continue
                ct = CONTRACT_TYPES.get(doc.document_type, {})
                extracted = db.query(FinanzExtractedData).filter(
                    FinanzExtractedData.document_id == doc.id
                ).all()
                fields = {ed.field_name: ed.field_value for ed in extracted}
                ws2.append([
                    ct.get("label", doc.document_type),
                    fields.get("gesellschaft", ""),
                    fields.get("beitrag", fields.get("beitrag_netto", "")),
                    doc.status,
                    doc.original_filename,
                    f"{doc.classification_confidence:.0%}" if doc.classification_confidence else "",
                ])

            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws2.column_dimensions[col].width = 25

            # --- Sheet 3: Detailfelder (with source references) ---
            ws3 = wb.create_sheet("Detailfelder")
            ws3.append([
                "Vertragsart", "Feld", "Wert", "Prioritaet",
                "Quelle (Dokument)", "Quelle (Seite)", "Confidence", "Verifiziert",
            ])
            style_header(ws3)

            for doc in docs:
                if not doc.document_type or doc.status != DocumentStatus.ANALYZED:
                    continue
                ct = CONTRACT_TYPES.get(doc.document_type, {})
                extracted = db.query(FinanzExtractedData).filter(
                    FinanzExtractedData.document_id == doc.id
                ).all()

                field_defs = get_fields_for_type(doc.document_type)
                field_map = {f["name"]: f for f in field_defs}

                for ed in extracted:
                    fdef = field_map.get(ed.field_name, {})
                    ws3.append([
                        ct.get("label", doc.document_type),
                        fdef.get("label", ed.field_name),
                        ed.field_value or "",
                        PRIORITY_LABELS.get(fdef.get("priority", ""), ""),
                        doc.original_filename,
                        f"Seite {ed.source_page}" if ed.source_page else "",
                        f"{ed.confidence:.0%}" if ed.confidence else "",
                        "Ja" if ed.verified else "Nein",
                    ])

            for i, w in enumerate([25, 30, 30, 12, 30, 12, 12, 12], 1):
                ws3.column_dimensions[chr(64 + i)].width = w

            # --- Sheet 4: Kosten ---
            ws4 = wb.create_sheet("Kosten")
            ws4.append(["Kategorie", "Vertragsart", "Beitrag (EUR)"])
            style_header(ws4)

            total_all = 0.0
            for cat_key, cat_def in CHECKLIST_CATEGORIES.items():
                cat_total = 0.0
                for type_key in cat_def["types"]:
                    if type_key not in {doc.document_type for doc in docs}:
                        continue
                    for doc in docs:
                        if doc.document_type != type_key:
                            continue
                        extracted = db.query(FinanzExtractedData).filter(
                            FinanzExtractedData.document_id == doc.id,
                            FinanzExtractedData.field_name.in_(
                                ["beitrag", "beitrag_netto", "sparbeitrag",
                                 "eigenbeitrag", "arbeitnehmerbeitrag", "sparrate"]
                            ),
                        ).all()
                        for ed in extracted:
                            if ed.field_value:
                                try:
                                    val = float(str(ed.field_value).replace(',', '.'))
                                    ct = CONTRACT_TYPES.get(type_key, {})
                                    ws4.append([
                                        cat_def["label"],
                                        ct.get("label", type_key),
                                        f"{val:.2f}",
                                    ])
                                    cat_total += val
                                except (ValueError, TypeError):
                                    pass
                total_all += cat_total

            ws4.append([])
            ws4.append(["GESAMT", "", f"{total_all:.2f}"])
            ws4['A' + str(ws4.max_row)].font = Font(bold=True)
            ws4['C' + str(ws4.max_row)].font = Font(bold=True)
            ws4.column_dimensions['A'].width = 25
            ws4.column_dimensions['B'].width = 35
            ws4.column_dimensions['C'].width = 15

            # --- Sheet 5: Optimierung ---
            ws5 = wb.create_sheet("Optimierung")
            ws5.append(["Kategorie", "Bewertung", "Zusammenfassung", "Details"])
            style_header(ws5)

            for sc in scorecards:
                if sc.is_overall:
                    continue
                label = cat_labels.get(sc.category, sc.category)
                ws5.append([
                    label,
                    rating_labels.get(sc.rating, sc.rating),
                    sc.assessment or "",
                    sc.details or "",
                ])

            ws5.append([])
            ws5.append(["Fehlende Pflichtfelder", "", "", ""])
            ws5['A' + str(ws5.max_row)].font = Font(bold=True)

            for doc in docs:
                if not doc.document_type or doc.status != DocumentStatus.ANALYZED:
                    continue
                ct = CONTRACT_TYPES.get(doc.document_type, {})
                extracted = db.query(FinanzExtractedData).filter(
                    FinanzExtractedData.document_id == doc.id
                ).all()
                field_names = {ed.field_name for ed in extracted if ed.field_value}
                field_defs = get_fields_for_type(doc.document_type)
                missing = [
                    f for f in field_defs
                    if f["priority"] == "muss" and f["name"] not in field_names
                ]
                for m in missing:
                    ws5.append([ct.get("label", doc.document_type), "MUSS", m["label"], "Fehlt"])

            for col in ['A', 'B', 'C', 'D']:
                ws5.column_dimensions[col].width = 30

            # --- Sheet 6: Rohdaten ---
            ws6 = wb.create_sheet("Rohdaten")
            ws6.append([
                "ID", "Document ID", "Feldname", "Wert", "Feldtyp",
                "Confidence", "Seite", "Quelltext", "Verifiziert",
                "Verifiziert von", "Verifiziert am",
            ])
            style_header(ws6)

            for doc in docs:
                extracted = db.query(FinanzExtractedData).filter(
                    FinanzExtractedData.document_id == doc.id
                ).all()
                for ed in extracted:
                    ws6.append([
                        ed.id,
                        ed.document_id,
                        ed.field_name,
                        ed.field_value or "",
                        ed.field_type,
                        ed.confidence,
                        ed.source_page,
                        (ed.source_text or "")[:200],
                        "Ja" if ed.verified else "Nein",
                        ed.verified_by or "",
                        ed.verified_at.strftime('%d.%m.%Y %H:%M') if ed.verified_at else "",
                    ])

            from openpyxl.utils import get_column_letter
            for i, w in enumerate([8, 12, 20, 30, 10, 10, 8, 50, 10, 15, 18], 1):
                ws6.column_dimensions[get_column_letter(i)].width = w

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer

        except ValueError:
            raise
        except Exception as e:
            logger.error("Excel export error for session %s: %s", session_id, e, exc_info=True)
            raise
        finally:
            db.close()

    def export_pdf(self, session_id: int) -> io.BytesIO:
        """
        Generate PDF export.

        Args:
            session_id: Session ID

        Returns:
            BytesIO buffer containing the PDF file
        """
        if not HAS_REPORTLAB:
            raise RuntimeError("reportlab nicht installiert — PDF-Export nicht verfügbar")
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak,
        )

        db = get_db_session()
        try:
            session = db.query(FinanzSession).filter(
                FinanzSession.id == session_id
            ).first()
            if session is None:
                raise ValueError(f"Session {session_id} not found")

            docs = db.query(FinanzDocument).filter(
                FinanzDocument.session_id == session_id,
            ).order_by(FinanzDocument.created_at).all()

            scorecards = db.query(FinanzScorecard).filter(
                FinanzScorecard.session_id == session_id
            ).all()

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, pagesize=A4,
                leftMargin=20*mm, rightMargin=20*mm,
                topMargin=25*mm, bottomMargin=20*mm,
            )

            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                'ZFATitle', parent=styles['Title'],
                fontSize=24, textColor=colors.HexColor('#294c5d'),
                spaceAfter=20,
            ))
            styles.add(ParagraphStyle(
                'ZFAHeading', parent=styles['Heading1'],
                fontSize=16, textColor=colors.HexColor('#294c5d'),
                spaceAfter=10,
            ))
            styles.add(ParagraphStyle(
                'ZFASubheading', parent=styles['Heading2'],
                fontSize=13, textColor=colors.HexColor('#207487'),
                spaceAfter=8,
            ))

            elements = []

            # --- Page 1: Cover ---
            elements.append(Spacer(1, 40*mm))
            elements.append(Paragraph("Finanzberatung", styles['ZFATitle']))
            elements.append(Paragraph(
                f"Analyse-Report fuer {session.customer_name}",
                styles['ZFAHeading'],
            ))
            elements.append(Spacer(1, 20*mm))

            info_data = [
                ["Kunde:", session.customer_name],
                ["Beratungsart:", session.session_type.capitalize()],
                ["Opener:", session.opener_username],
                ["Closer:", session.closer_username or "Nicht zugewiesen"],
                ["Termin:", session.appointment_date.strftime('%d.%m.%Y') if session.appointment_date else "-"],
                ["Erstellt am:", datetime.now().strftime('%d.%m.%Y %H:%M')],
            ]
            info_table = Table(info_data, colWidths=[40*mm, 100*mm])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#77726d')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(info_table)
            elements.append(PageBreak())

            # --- Page 2: Ampel Overview ---
            elements.append(Paragraph("Ampel-Uebersicht", styles['ZFAHeading']))
            elements.append(Spacer(1, 5*mm))

            cat_labels = {
                ScorecardCategory.ALTERSVORSORGE: "Altersvorsorge",
                ScorecardCategory.ABSICHERUNG: "Absicherung",
                ScorecardCategory.VERMOEGEN_KOSTEN: "Vermoegen & Kosten",
                ScorecardCategory.STEUEROPTIMIERUNG: "Steueroptimierung",
            }
            rating_colors = {
                TrafficLight.GREEN: colors.HexColor('#22c55e'),
                TrafficLight.YELLOW: colors.HexColor('#eab308'),
                TrafficLight.RED: colors.HexColor('#ef4444'),
            }

            sc_data = [["Kategorie", "Bewertung", "Zusammenfassung"]]
            for sc in scorecards:
                if sc.is_overall:
                    continue
                label = cat_labels.get(sc.category, sc.category)
                rating_text = {
                    TrafficLight.GREEN: "GRUEN",
                    TrafficLight.YELLOW: "GELB",
                    TrafficLight.RED: "ROT",
                }.get(sc.rating, sc.rating)
                sc_data.append([label, rating_text, sc.assessment or ""])

            if len(sc_data) > 1:
                sc_table = Table(sc_data, colWidths=[40*mm, 25*mm, 95*mm])
                table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#294c5d')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                ]
                # Color code rating cells
                for i, sc in enumerate(scorecards):
                    if sc.is_overall:
                        continue
                    row = i + 1
                    if row < len(sc_data):
                        bg = rating_colors.get(sc.rating)
                        if bg:
                            table_style.append(('BACKGROUND', (1, row), (1, row), bg))
                            table_style.append(('TEXTCOLOR', (1, row), (1, row), colors.white))

                sc_table.setStyle(TableStyle(table_style))
                elements.append(sc_table)

            # Overall
            overall = next((sc for sc in scorecards if sc.is_overall), None)
            if overall:
                elements.append(Spacer(1, 10*mm))
                elements.append(Paragraph(
                    f"<b>Gesamtbewertung:</b> {overall.assessment or ''}",
                    styles['Normal'],
                ))

            elements.append(PageBreak())

            # --- Page 3+: Contract Details ---
            elements.append(Paragraph("Vertragsdetails", styles['ZFAHeading']))
            elements.append(Spacer(1, 5*mm))

            for doc_obj in docs:
                if not doc_obj.document_type or doc_obj.status != DocumentStatus.ANALYZED:
                    continue

                ct = CONTRACT_TYPES.get(doc_obj.document_type, {})
                elements.append(Paragraph(
                    ct.get("label", doc_obj.document_type),
                    styles['ZFASubheading'],
                ))
                elements.append(Paragraph(
                    f"<i>Dokument: {doc_obj.original_filename}</i>",
                    styles['Normal'],
                ))

                extracted = db.query(FinanzExtractedData).filter(
                    FinanzExtractedData.document_id == doc_obj.id
                ).all()

                if extracted:
                    field_defs = get_fields_for_type(doc_obj.document_type)
                    field_map = {f["name"]: f for f in field_defs}
                    detail_data = [["Feld", "Wert", "Quelle", "Status"]]
                    for ed in extracted:
                        fdef = field_map.get(ed.field_name, {})
                        source = f"S. {ed.source_page}" if ed.source_page else ""
                        status = "Verifiziert" if ed.verified else PRIORITY_LABELS.get(fdef.get("priority", ""), "")
                        detail_data.append([
                            fdef.get("label", ed.field_name),
                            str(ed.field_value or "")[:60],
                            source,
                            status,
                        ])

                    detail_table = Table(detail_data, colWidths=[45*mm, 55*mm, 20*mm, 30*mm])
                    detail_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#207487')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ]))
                    elements.append(detail_table)

                elements.append(Spacer(1, 8*mm))

            # --- Last section: Handlungsbedarf ---
            elements.append(PageBreak())
            elements.append(Paragraph("Handlungsbedarf", styles['ZFAHeading']))
            elements.append(Spacer(1, 5*mm))

            for doc_obj in docs:
                if not doc_obj.document_type or doc_obj.status != DocumentStatus.ANALYZED:
                    continue
                ct = CONTRACT_TYPES.get(doc_obj.document_type, {})
                extracted = db.query(FinanzExtractedData).filter(
                    FinanzExtractedData.document_id == doc_obj.id
                ).all()
                field_names = {ed.field_name for ed in extracted if ed.field_value}
                field_defs = get_fields_for_type(doc_obj.document_type)
                missing = [
                    f for f in field_defs
                    if f["priority"] == "muss" and f["name"] not in field_names
                ]
                if missing:
                    elements.append(Paragraph(
                        f"<b>{ct.get('label', doc_obj.document_type)}</b> - Fehlende Pflichtfelder:",
                        styles['Normal'],
                    ))
                    for m in missing:
                        elements.append(Paragraph(
                            f"&bull; {m['label']}",
                            styles['Normal'],
                        ))
                    elements.append(Spacer(1, 4*mm))

            # Scorecard recommendations
            elements.append(Spacer(1, 5*mm))
            elements.append(Paragraph("Empfehlungen", styles['ZFASubheading']))
            for sc in scorecards:
                if sc.is_overall or not sc.details:
                    continue
                label = cat_labels.get(sc.category, sc.category)
                elements.append(Paragraph(
                    f"<b>{label}:</b> {sc.details}",
                    styles['Normal'],
                ))
                elements.append(Spacer(1, 3*mm))

            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            return buffer

        except ValueError:
            raise
        except Exception as e:
            logger.error("PDF export error for session %s: %s", session_id, e, exc_info=True)
            raise
        finally:
            db.close()
