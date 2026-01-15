# -*- coding: utf-8 -*-
"""
Admin Tracking Export Routes
Excel und CSV Exports für Tracking-Analytics
"""

import csv
import io
from datetime import datetime

import pytz
from flask import jsonify, request, Response

from app.config.base import slot_config
from app.core.extensions import tracking_system
from app.utils.decorators import require_admin
from app.routes.admin import admin_bp

TZ = pytz.timezone(slot_config.TIMEZONE)


@admin_bp.route("/tracking/export/csv")
@require_admin
def export_tracking_csv():
    """
    Export Tracking-Daten als CSV

    Query Parameters:
        start_date: Start-Datum YYYY-MM-DD (default: 2025-09-01)
        end_date: End-Datum YYYY-MM-DD (default: heute)
        type: 'daily' | 'summary' (default: daily)
    """
    try:
        if not tracking_system:
            return jsonify({"success": False, "error": "Tracking-System nicht verfügbar"}), 503

        start_date = request.args.get("start_date", "2025-09-01")
        end_date = request.args.get("end_date", str(datetime.now(TZ).date()))
        export_type = request.args.get("type", "daily")

        # Hole Daten
        stats = tracking_system.get_stats_for_period(start_date, end_date)

        # Erstelle CSV
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')

        if export_type == "summary":
            # Zusammenfassung
            writer.writerow(["Tracking Analytics Export - Zusammenfassung"])
            writer.writerow([f"Zeitraum: {start_date} bis {end_date}"])
            writer.writerow([])
            writer.writerow(["Metrik", "Wert"])
            writer.writerow(["Getrackte Tage", stats.get("days_tracked", 0)])
            writer.writerow(["Gesamte Termine", stats.get("total_slots", 0)])
            writer.writerow(["Erschienen", stats.get("completed", 0)])
            writer.writerow(["Nicht erschienen", stats.get("no_shows", 0)])
            writer.writerow(["Abgesagt", stats.get("cancelled", 0)])
            writer.writerow(["Verschoben", stats.get("rescheduled", 0)])
            writer.writerow(["Auftauchquote", f"{stats.get('appearance_rate', 0)}%"])
            writer.writerow(["No-Show-Rate", f"{stats.get('no_show_rate', 0)}%"])
        else:
            # Tagesstatistiken
            writer.writerow([
                "Datum", "Wochentag", "Termine gesamt", "Erschienen",
                "Nicht erschienen", "Abgesagt", "Verschoben", "Auftauchquote"
            ])

            for day in stats.get("daily_data", []):
                writer.writerow([
                    day.get("date", ""),
                    day.get("weekday", ""),
                    day.get("total_slots", 0),
                    day.get("completed", 0),
                    day.get("no_shows", 0),
                    day.get("cancelled", 0),
                    day.get("rescheduled", 0),
                    f"{day.get('appearance_rate', 0)}%"
                ])

        # Response erstellen
        output.seek(0)
        filename = f"tracking_export_{start_date}_{end_date}.csv"

        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "text/csv; charset=utf-8"
            }
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route("/tracking/export/excel")
@require_admin
def export_tracking_excel():
    """
    Export Tracking-Daten als Excel (4 Sheets)

    Query Parameters:
        start_date: Start-Datum YYYY-MM-DD (default: 2025-09-01)
        end_date: End-Datum YYYY-MM-DD (default: heute)
    """
    try:
        # Prüfe ob openpyxl verfügbar ist
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({
                "success": False,
                "error": "Excel-Export nicht verfügbar. Bitte openpyxl installieren."
            }), 503

        if not tracking_system:
            return jsonify({"success": False, "error": "Tracking-System nicht verfügbar"}), 503

        start_date = request.args.get("start_date", "2025-09-01")
        end_date = request.args.get("end_date", str(datetime.now(TZ).date()))

        # Hole Daten
        stats = tracking_system.get_stats_for_period(start_date, end_date)
        consultant_perf = tracking_system.get_consultant_performance(start_date, end_date)

        # Combined Ranking
        from app.services.consultant_ranking import consultant_ranking_service
        ranking_data = consultant_ranking_service.get_ranking_summary(start_date, end_date)

        # Erstelle Workbook
        wb = Workbook()

        # Styles definieren
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="294C5D", end_color="294C5D", fill_type="solid")
        good_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        bad_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ========== Sheet 1: Übersicht ==========
        ws1 = wb.active
        ws1.title = "Übersicht"

        ws1["A1"] = "Tracking Analytics Export"
        ws1["A1"].font = Font(bold=True, size=16)
        ws1["A2"] = f"Zeitraum: {start_date} bis {end_date}"
        ws1["A3"] = f"Exportiert: {datetime.now(TZ).strftime('%d.%m.%Y %H:%M')}"

        # Übersichts-Tabelle
        overview_start = 5
        overview_data = [
            ["Metrik", "Wert"],
            ["Getrackte Tage", stats.get("days_tracked", 0)],
            ["Gesamte Termine", stats.get("total_slots", 0)],
            ["Erschienen", stats.get("completed", 0)],
            ["Nicht erschienen", stats.get("no_shows", 0)],
            ["Abgesagt", stats.get("cancelled", 0)],
            ["Verschoben", stats.get("rescheduled", 0)],
            ["Auftauchquote", f"{stats.get('appearance_rate', 0)}%"],
            ["No-Show-Rate", f"{stats.get('no_show_rate', 0)}%"]
        ]

        for row_idx, row_data in enumerate(overview_data, overview_start):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx == overview_start:
                    cell.font = header_font
                    cell.fill = header_fill

        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 15

        # ========== Sheet 2: Tagesstatistik ==========
        ws2 = wb.create_sheet("Tagesstatistik")

        headers = ["Datum", "Wochentag", "Termine", "Erschienen", "No-Shows",
                   "Abgesagt", "Verschoben", "Auftauchquote"]

        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, day in enumerate(stats.get("daily_data", []), 2):
            appearance_rate = day.get("appearance_rate", 0)

            data = [
                day.get("date", ""),
                day.get("weekday", ""),
                day.get("total_slots", 0),
                day.get("completed", 0),
                day.get("no_shows", 0),
                day.get("cancelled", 0),
                day.get("rescheduled", 0),
                f"{appearance_rate}%"
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Färbe Auftauchquote
                if col_idx == 8:
                    if appearance_rate >= 80:
                        cell.fill = good_fill
                    elif appearance_rate >= 60:
                        cell.fill = warning_fill
                    else:
                        cell.fill = bad_fill

        # Spaltenbreiten
        for col in range(1, 9):
            ws2.column_dimensions[get_column_letter(col)].width = 15

        # ========== Sheet 3: Berater Show-Rates ==========
        ws3 = wb.create_sheet("Berater Show-Rates")

        headers = ["Berater", "Termine", "Erschienen", "No-Shows",
                   "Abgesagt", "Verschoben", "Auftauchquote"]

        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row_idx = 2
        for consultant, perf in sorted(consultant_perf.items(),
                                       key=lambda x: x[1].get("appearance_rate", 0),
                                       reverse=True):
            appearance_rate = perf.get("appearance_rate", 0)

            data = [
                consultant,
                perf.get("total_slots", 0),
                perf.get("completed", 0),
                perf.get("no_shows", 0),
                perf.get("cancelled", 0),
                perf.get("rescheduled", 0),
                f"{appearance_rate}%"
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                if col_idx == 7:
                    if appearance_rate >= 80:
                        cell.fill = good_fill
                    elif appearance_rate >= 60:
                        cell.fill = warning_fill
                    else:
                        cell.fill = bad_fill

            row_idx += 1

        for col in range(1, 8):
            ws3.column_dimensions[get_column_letter(col)].width = 15

        # ========== Sheet 4: Combined Ranking ==========
        ws4 = wb.create_sheet("Combined Ranking")

        headers = ["Rang", "Berater", "Combined Score", "Kategorie",
                   "Show-Rate", "Tel. Achievement", "Aktivitäten"]

        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row_idx = 2
        for r in ranking_data.get("rankings", []):
            category = r.get("category", "")
            data = [
                r.get("rank", ""),
                r.get("name", ""),
                r.get("combined_score", 0),
                category.upper(),
                f"{r.get('show_rate', 0)}%",
                f"{r.get('telefonie_achievement', 0)}%",
                r.get("total_activities", 0)
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                if col_idx == 4:
                    if category == "high":
                        cell.fill = good_fill
                    elif category == "medium":
                        cell.fill = warning_fill
                    else:
                        cell.fill = bad_fill

            row_idx += 1

        for col in range(1, 8):
            ws4.column_dimensions[get_column_letter(col)].width = 18

        # Speichere in BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"tracking_analytics_{start_date}_{end_date}.xlsx"

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
